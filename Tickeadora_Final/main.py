import sys
import openpyxl
import csv
import os
import pandas as pd
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QWidget, QPushButton, QLabel, QVBoxLayout, QHBoxLayout,
    QFileDialog, QGridLayout, QMessageBox, QScrollArea, QLineEdit, QDialog,
    QListWidget, QButtonGroup
)
from PyQt5.QtGui import QFont, QColor, QPalette
from PyQt5.QtCore import Qt, QTimer
from escpos.printer import Usb, Network
from collections import defaultdict

class ComanderaApp(QWidget):
    def __init__(self):
        super().__init__()
        self.inicializar_impresora()
        self.setWindowTitle("Comandera Visual C.A.S.")
        self.setGeometry(100, 100, 1000, 700)

        self.ventas = []            # Lista de ventas realizadas
        self.productos = {}         # Diccionario {nombre: precio}
        self.ultima_parcial = 0     # Índice desde donde tomar la próxima venta parcial
        self.suma_productos = 0.0   # Suma de productos impresos para cada venta 
        self.total_efectivo = 0.0   # Total de ventas en efectivo
        self.total_qr = 0.0         # Total de ventas por QR
        self.parcial_efectivo = 0.0 # Suma parcial de efectivo
        self.parcial_qr = 0.0       # Suma parcial de QR

        self.init_ui()

    def init_ui(self):
        # Con QVBoxLayout se organiza verticalmente los widgets
        layout = QVBoxLayout()

        self.label_info = QLabel("Cargá productos desde un archivo Excel")
        self.label_info.setFont(QFont("Arial", 16))
        self.label_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label_info)

        boton_manual = QPushButton("📋 Cargar Manualmente Productos")
        boton_manual.setFont(QFont("Arial", 14))
        boton_manual.clicked.connect(self.abrir_carga_manual)
        layout.addWidget(boton_manual)

        boton_cargar = QPushButton("📂 Cargar Productos (Excel)")
        boton_cargar.setFont(QFont("Arial", 14))
        boton_cargar.clicked.connect(self.cargar_productos)
        layout.addWidget(boton_cargar)

        suma_layout = QHBoxLayout()

        self.suma_label = QLineEdit("$ 0.00")
        self.suma_label.setReadOnly(True)
        self.suma_label.setFont(QFont("Arial", 16))
        self.suma_label.setStyleSheet("background-color: lightblue; border: 1px solid black;")
        self.suma_label.setAlignment(Qt.AlignCenter)
        self.suma_label.setFixedWidth(960)
        #suma_layout.addWidget(QLabel("Suma productos impresos:"))
        suma_layout.addWidget(self.suma_label)

        btn_reiniciar_suma = QPushButton("Nueva Venta")
        btn_reiniciar_suma.setFont(QFont("Arial", 15))
        btn_reiniciar_suma.clicked.connect(self.reiniciar_suma_productos)
        suma_layout.addWidget(btn_reiniciar_suma)
        layout.addLayout(suma_layout)

        self.btn_Efectivo = QPushButton("Efectivo")
        self.btn_Efectivo.setFont(QFont("Arial", 15))
        self.btn_Efectivo.setCheckable(True)
        suma_layout.addWidget(self.btn_Efectivo)

        self.btn_QR = QPushButton("QR")
        self.btn_QR.setFont(QFont("Arial", 15))
        self.btn_QR.setCheckable(True)
        suma_layout.addWidget(self.btn_QR)

        self.grupo_botones = QButtonGroup(self)
        self.grupo_botones.setExclusive(True)
        self.grupo_botones.addButton(self.btn_Efectivo)
        self.grupo_botones.addButton(self.btn_QR)

        self.grid_widget = QWidget()
        self.grid_layout = QGridLayout()
        self.grid_widget.setLayout(self.grid_layout)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.grid_widget)
        layout.addWidget(scroll)

        # Con QHBoxLayout se organizan horizontalmente los botones de funciones
        botones_func = QHBoxLayout()

        self.btn_ticket_parcial = QPushButton("🧾 Ticket Parcial")
        self.btn_ticket_parcial.setFont(QFont("Arial", 14))
        self.btn_ticket_parcial.clicked.connect(self.ticket_parcial)
        botones_func.addWidget(self.btn_ticket_parcial)

        self.btn_ticket_total = QPushButton("🧾 Ticket Total")
        self.btn_ticket_total.setFont(QFont("Arial", 14))
        self.btn_ticket_total.clicked.connect(self.ticket_total)
        botones_func.addWidget(self.btn_ticket_total)

        self.btn_reiniciar = QPushButton("🔄 Reiniciar Todo")
        self.btn_reiniciar.setFont(QFont("Arial", 14))
        self.btn_reiniciar.clicked.connect(self.reiniciar_ventas)
        botones_func.addWidget(self.btn_reiniciar)

        self.btn_salir = QPushButton("🚪 Salir")
        self.btn_salir.setFont(QFont("Arial", 14))
        self.btn_salir.clicked.connect(self.close)
        botones_func.addWidget(self.btn_salir)

        layout.addLayout(botones_func)
        self.setLayout(layout)

    def inicializar_impresora(self):
        try:
            #self.impresora = Network("192.168.0.241")
            self.impresora = Usb(0x1fc9, 0x2016) # ID de la impresora
            #self.impresora.set(width=4, height=4)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo conectar a la impresora:\n{str(e)}")
            self.impresora = None

    def cargar_productos(self):
        archivo, _ = QFileDialog.getOpenFileName(self, "Seleccionar archivo Excel", "", "Excel Files (*.xlsx)")
        if archivo:
            try:
                wb = openpyxl.load_workbook(archivo)
                hoja = wb.active
                self.productos.clear()
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    nombre, precio = fila
                    if nombre and precio:
                        self.productos[nombre] = float(precio)
                self.actualizar_botones()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo cargar el archivo:\n{str(e)}")

    def actualizar_botones(self):
        for i in reversed(range(self.grid_layout.count())):
            widget = self.grid_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        fila, col = 0, 0
        for idx, (nombre, precio) in enumerate(self.productos.items()):
            btn = QPushButton(f"{nombre}\n${precio:.2f}")
            btn.setMinimumSize(200, 200)
            btn.setFont(QFont("Arial", 18))
            btn.setStyleSheet("background-color: lightblue; border-radius: 10px;")
            btn.clicked.connect(lambda _, n=nombre, b=btn: self.registrar_venta(n, b))
            self.grid_layout.addWidget(btn, fila, col)
            col += 1
            if col > 3:
                col = 0
                fila += 1

    def imprimir_producto(self, nombre, precio):
        if self.impresora:
            self.impresora.set(align='center')
            self.impresora._raw(b'\x1D\x21\x11')
            self.impresora.text("======== TICKET ========\n")
            self.impresora.set(align='center', height=2, width=2)
            self.impresora.text(f"{nombre}\n")
            self.impresora.set(align='center', height=2, width=2)
            self.impresora.text(f"Precio: ${precio:.2f}\n")
            self.impresora.set(align='center', height=2, width=2)
            self.impresora.text("========================\n\n")
            
            self.impresora.cut()

    def registrar_venta(self, nombre, boton):
        ahora = datetime.now()
        precio = self.productos[nombre]
        self.ventas.append((nombre, precio, None))
        print(f"Producto vendido: {nombre} - ${precio:.2f} - Hora: {ahora.strftime('%H:%M:%S')}")
        self.imprimir_producto(nombre, precio)

        # Actualiza suma de productos impresos
        self.suma_productos += precio
        self.suma_label.setText(f"$ {self.suma_productos:.2f}")

        # Cambia color para feedback visual
        original_style = boton.styleSheet()
        boton.setStyleSheet("background-color: green; color: white; border-radius: 10px;")
        QTimer.singleShot(200, lambda: boton.setStyleSheet(original_style))

    def reiniciar_suma_productos(self):
        monto = self.suma_productos
        if monto > 0:
            if self.btn_Efectivo.isChecked():
                metodo = "Efectivo"
                self.total_efectivo += monto
                self.parcial_efectivo += monto
            elif self.btn_QR.isChecked():
                metodo = "QR"
                self.total_qr += monto
                self.parcial_qr += monto
            else:
                metodo = "Efectivo"  # Por defecto
                self.total_efectivo += monto
                self.parcial_efectivo += monto

            nuevas_ventas = []
            for nombre, precio, m in self.ventas:
                if m is None:
                    nuevas_ventas.append((nombre, precio, metodo))
                else:
                    nuevas_ventas.append((nombre, precio, m))
            self.ventas = nuevas_ventas

        # Reset suma
        self.suma_productos = 0.0
        self.suma_label.setText("$ 0.00")

    """
    def ticket_parcial(self):
        nuevas_ventas = self.ventas[self.ultima_parcial:]
        total = sum(p for _, p in nuevas_ventas)
        resumen = "\n".join(f"{n}: ${p:.2f}" for n, p in nuevas_ventas)
        print("\n--- Ticket Parcial ---")
        print(resumen)
        print(f"\n🧾 PARCIAL: ${total:.2f}")
        print("----------------------\n")
        self.imprimir_ticket("Ticket Parcial", nuevas_ventas)
        self.ultima_parcial = len(self.ventas)
    """

    def registrar_ventas_pendientes(self):
        # Detecta si hay ventas con método None y las asigna al método actual
        metodo = None
        if self.btn_Efectivo.isChecked():
            metodo = "Efectivo"
        elif self.btn_QR.isChecked():
            metodo = "QR"
        else:
            metodo = "Efectivo"  # Por defecto

        nuevas_ventas = []
        for nombre, precio, m in self.ventas:
            if m is None:
                nuevas_ventas.append((nombre, precio, metodo))
                # Suma a los acumulados
                if metodo == "Efectivo":
                    self.total_efectivo += precio
                    self.parcial_efectivo += precio
                else:
                    self.total_qr += precio
                    self.parcial_qr += precio
            else:
                nuevas_ventas.append((nombre, precio, m))
        self.ventas = nuevas_ventas
        # Resetea suma_productos y suma_label
        self.suma_productos = 0.0
        self.suma_label.setText("$ 0.00")

    def ticket_parcial(self):
        self.registrar_ventas_pendientes()
        nuevas_ventas = self.ventas[self.ultima_parcial:]
        if not nuevas_ventas:
            print("No hay nuevas ventas.")
            return

        resumen = defaultdict(lambda: {"Efectivo": [0, 0.0], "QR": [0, 0.0]})
        for nombre, precio, metodo in nuevas_ventas:
            resumen[nombre][metodo][0] += 1
            resumen[nombre][metodo][1] += precio

        ahora = datetime.now()
        fecha_hora = ahora.strftime("%d/%m/%Y %H:%M:%S")

        total = 0
        print("\n--- Ticket Parcial ---")
        for nombre, pagos in resumen.items():
            efectivo_cant, efectivo_sub = pagos["Efectivo"]
            qr_cant, qr_sub = pagos["QR"]
            print(f"{nombre:<20} Efectivo: x{efectivo_cant:<2} ${efectivo_sub:>7.2f} | QR: x{qr_cant:<2} ${qr_sub:>7.2f}")
            total += efectivo_sub + qr_sub
        print(f"\nTOTAL PARCIAL: ${total:.2f}")
        print(f"Total Efectivo: ${self.parcial_efectivo:.2f}")
        print(f"Total QR: ${self.parcial_qr:.2f}")
        print("-----------------------\n")

        self.ultima_parcial = len(self.ventas)

        if self.impresora:
            self.impresora.text("------ PARCIAL ------\n")
            self.impresora.text(f"{fecha_hora}\n")
            self.impresora.text("--- EFECTIVO ---\n")
            for nombre, pagos in resumen.items():
                efectivo_cant, efectivo_sub = pagos["Efectivo"]
                #qr_cant, qr_sub = pagos["QR"]
                #linea_efectivo = f"{nombre:<20} Efectivo: x{efectivo_cant:<2} ${efectivo_sub:>7.2f} | QR: x{qr_cant:<2} ${qr_sub:>7.2f}\n"
                linea_efectivo = f"{nombre:<20} x{efectivo_cant:<2} ${efectivo_sub:>7.2f}\n"
                self.impresora.text(linea_efectivo)
            self.impresora.text("--- QR ---\n")
            for nombre, pagos in resumen.items():
                #efectivo_cant, efectivo_sub = pagos["Efectivo"]
                qr_cant, qr_sub = pagos["QR"]
                #linea_efectivo = f"{nombre:<20} Efectivo: x{efectivo_cant:<2} ${efectivo_sub:>7.2f} | QR: x{qr_cant:<2} ${qr_sub:>7.2f}\n"
                linea_QR = f"{nombre:<20} x{qr_cant:<2} ${qr_sub:>7.2f}\n"
                self.impresora.text(linea_QR)
            self.impresora.text(f"\nTOTAL: ${total:.2f}\n")
            self.impresora.text(f"Efectivo: ${self.parcial_efectivo:.2f}\n")
            self.impresora.text(f"QR: ${self.parcial_qr:.2f}\n")
            self.impresora.text("----------------------\n\n")
            self.impresora.cut()

        # Reiniciar los acumulados parciales para el próximo ticket parcial
        self.parcial_efectivo = 0.0
        self.parcial_qr = 0.0

    """ 
    def ticket_total(self):
        total = sum(p for _, p in self.ventas)
        resumen = "\n".join(f"{n}: ${p:.2f}" for n, p in self.ventas)
        print("\n--- Ticket Total ---")
        print(resumen)
        print(f"\n🧾 TOTAL: ${total:.2f}")
        print("---------------------\n")
        self.imprimir_ticket("Ticket Total", self.ventas) 
    """

    def ticket_total(self):
        self.registrar_ventas_pendientes()
        if not self.ventas:
            print("No hay ventas.")
            return

        resumen = defaultdict(lambda: {"Efectivo": [0, 0.0], "QR": [0, 0.0]})
        for nombre, precio, metodo in self.ventas:
            resumen[nombre][metodo][0] += 1
            resumen[nombre][metodo][1] += precio

        ahora = datetime.now()
        fecha_hora = ahora.strftime("%d/%m/%Y %H:%M:%S")

        total = 0
        print("\n--- Ticket Total ---")
        for nombre, pagos in resumen.items():
            efectivo_cant, efectivo_sub = pagos["Efectivo"]
            qr_cant, qr_sub = pagos["QR"]
            print(f"{nombre:<20} Efectivo: x{efectivo_cant:<2} ${efectivo_sub:>7.2f} | QR: x{qr_cant:<2} ${qr_sub:>7.2f}")
            total += efectivo_sub + qr_sub
        print(f"\nTOTAL: ${total:.2f}")
        print(f"Total Efectivo: ${self.total_efectivo:.2f}")
        print(f"Total QR: ${self.total_qr:.2f}")
        print("---------------------\n")

        if self.impresora:
            self.impresora.text("======= TOTAL =======\n")
            self.impresora.text(f"{fecha_hora}\n")

            self.impresora.text("--- EFECTIVO ---\n")
            for nombre, pagos in resumen.items():
                efectivo_cant, efectivo_sub = pagos["Efectivo"]
                #qr_cant, qr_sub = pagos["QR"]
                #linea_efectivo = f"{nombre:<20} Efectivo: x{efectivo_cant:<2} ${efectivo_sub:>7.2f} | QR: x{qr_cant:<2} ${qr_sub:>7.2f}\n"
                linea_efectivo = f"{nombre:<20} x{efectivo_cant:<2} ${efectivo_sub:>7.2f}\n"
                self.impresora.text(linea_efectivo)
            self.impresora.text("--- QR ---\n")
            for nombre, pagos in resumen.items():
                #efectivo_cant, efectivo_sub = pagos["Efectivo"]
                qr_cant, qr_sub = pagos["QR"]
                #linea_efectivo = f"{nombre:<20} Efectivo: x{efectivo_cant:<2} ${efectivo_sub:>7.2f} | QR: x{qr_cant:<2} ${qr_sub:>7.2f}\n"
                linea_QR = f"{nombre:<20} x{qr_cant:<2} ${qr_sub:>7.2f}\n"
                self.impresora.text(linea_QR)
            
            self.impresora.text(f"\nTOTAL: ${total:.2f}\n")
            self.impresora.text(f"Efectivo: ${self.total_efectivo:.2f}\n")
            self.impresora.text(f"QR: ${self.total_qr:.2f}\n")
            self.impresora.text("=====================\n\n")
            self.impresora.cut()

    def reiniciar_ventas(self):
        confirmar = QMessageBox.question(self, "Confirmar", "¿Seguro que querés borrar todos los datos de ventas?",
                                         QMessageBox.Yes | QMessageBox.No)
        if confirmar == QMessageBox.Yes:
            self.guardar_backup()
            self.ventas.clear()
            self.ultima_parcial = 0
            self.total_efectivo = 0.0
            self.total_qr = 0.0
            print("✅ Ventas reiniciadas correctamente.")

    def guardar_backup(self):
        if not self.ventas:
            return
        backup_path = os.path.join(os.getcwd(), "ventas_backup.csv")
        with open(backup_path, mode='w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Producto", "Precio"])
            writer.writerows(self.ventas)
        print(f"💾 Backup guardado en: {backup_path}")

    def abrir_carga_manual(self):
        dialogo = CargaManualDialog(self)
        dialogo.exec_()


class CargaManualDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Carga Manual de Productos")
        self.setGeometry(100, 100, 1200, 900)
        self.lista_productos = QListWidget()
        self.cargar_productos_existentes()

        layout = QVBoxLayout()

        form_layout = QHBoxLayout()
        self.nombre_input = QLineEdit()
        self.nombre_input.setPlaceholderText("Nombre del producto")
        self.nombre_input.setFont(QFont("Arial", 12))
        self.precio_input = QLineEdit()
        self.precio_input.setPlaceholderText("Precio")
        self.precio_input.setFont(QFont("Arial", 12))
        form_layout.addWidget(self.nombre_input)
        form_layout.addWidget(self.precio_input)
        layout.addLayout(form_layout)

        btn_agregar = QPushButton("➕ Agregar Producto")
        btn_agregar.setFont(QFont("Arial", 14))
        btn_agregar.clicked.connect(self.agregar_producto)
        layout.addWidget(btn_agregar)

        boton_borrar = QPushButton("🗑 Borrar Todos los Productos")
        boton_borrar.setFont(QFont("Arial", 14))
        boton_borrar.clicked.connect(self.borrar_productos)
        layout.addWidget(boton_borrar)

        layout_modificar = QHBoxLayout()    

        btn_modificar = QPushButton("✏️ Modificar Producto Sleccionado")
        btn_modificar.setFont(QFont("Arial", 14))
        btn_modificar.clicked.connect(self.modificar_producto)
        layout_modificar.addWidget(btn_modificar)

        btn_borrar_modificado = QPushButton("🗑 Borrar Producto Sleccionado")
        btn_borrar_modificado.setFont(QFont("Arial", 14))
        btn_borrar_modificado.clicked.connect(self.borrar_producto_seleccionado)
        layout_modificar.addWidget(btn_borrar_modificado)

        layout.addLayout(layout_modificar)

        self.label_lista = QLabel("Productos cargados:\n")
        self.label_lista.setFont(QFont("Arial", 16))
        self.lista_productos.itemClicked.connect(self.cargar_producto_en_campos)
        layout.addWidget(self.label_lista)

        layout.addWidget(self.lista_productos)

        btn_guardar = QPushButton("💾 Guardar como Excel")
        btn_guardar.setFont(QFont("Arial", 14))
        btn_guardar.clicked.connect(self.guardar_excel)
        layout.addWidget(btn_guardar)

        self.setLayout(layout)

    def agregar_producto(self):
        nombre = self.nombre_input.text().strip()
        try:
            precio = float(self.precio_input.text())
        except ValueError:
            QMessageBox.warning(self, "Error", "Precio inválido.")
            return

        if not nombre:
            QMessageBox.warning(self, "Error", "Nombre vacío.")
            return

        #self.lista_productos.append((nombre, precio))
        #self.label_lista.setText(self.label_lista.text() + f"- {nombre}: ${precio:.2f}\n")
        self.lista_productos.addItem(f"{nombre} - ${precio:.2f}")
        self.label_lista.setStyleSheet("font-size: 30px; font-family: Arial;")
        self.nombre_input.clear()
        self.precio_input.clear()

    def guardar_excel(self):
        if not self.lista_productos:
            QMessageBox.warning(self, "Error", "No hay productos cargados.")
            return

        archivo, _ = QFileDialog.getSaveFileName(self, "Guardar archivo Excel", "", "Excel Files (*.xlsx)")
        if archivo:
            if not archivo.endswith(".xlsx"):
                archivo += ".xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Producto", "Precio"])
            for i in range(self.lista_productos.count()):
                item_text = self.lista_productos.item(i).text()  # Ej: "Producto1 - $23.50"
                nombre, precio_str = item_text.split(" - $")
                precio = float(precio_str)
                ws.append([nombre, precio])

            wb.save(archivo)
            QMessageBox.information(self, "Éxito", f"Archivo guardado en:\n{archivo}")
            self.close()
    
    def cargar_productos_existentes(self):
        archivo = "productos.xlsx"
        if not os.path.exists(archivo):
            return

        wb = openpyxl.load_workbook(archivo)
        hoja = wb.active
        #self.productos.clear()
        self.lista_productos.clear()

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            nombre, precio = fila
            if nombre and precio is not None:
                #self.lista_productos.append({"Nombre": nombre, "Precio": float(precio)})
                self.lista_productos.addItem(f"{nombre} - ${precio:.2f}")


    def actualizar_lista_productos(self):
        self.lista_productos.clear()
        for producto in self.lista_productos:
            nombre = producto["Nombre"]
            precio = producto["Precio"]
            self.lista_productos.addItem(f"{nombre} - ${precio:.2f}")
            self.label_lista.setText(self.label_lista.text() + f"- {nombre}: ${precio:.2f}\n")
            self.label_lista.setStyleSheet("font-size: 30px; font-family: Arial;")

    def cargar_producto_en_campos(self, item):
        texto = item.text()  # Ejemplo: "Hamburguesa - $4000.00"
        if " - $" in texto:
            nombre, precio_str = texto.split(" - $")
            self.nombre_input.setText(nombre)
            self.precio_input.setText(precio_str)

    def modificar_producto(self):
        nombre = self.nombre_input.text().strip()
        try:
            precio = float(self.precio_input.text())
        except ValueError:
            QMessageBox.warning(self, "Error", "Precio inválido.")
            return

        if not nombre:
            QMessageBox.warning(self, "Error", "Nombre vacío.")
            return

        item = self.lista_productos.currentItem()
        if item:
            item.setText(f"{nombre} - ${precio:.2f}")
            self.nombre_input.clear()
            self.precio_input.clear()

    def borrar_producto_seleccionado(self):
        item = self.lista_productos.currentItem()
        if item:
            fila = self.lista_productos.row(item)
            self.lista_productos.takeItem(fila)
            self.nombre_input.clear()
            self.precio_input.clear()  

    # Borra por completo el excel de productos
    #def borrar_productos(self):
    #    self.lista_productos = []
    #    self.actualizar_lista_productos()
    #    archivo = "productos.xlsx"
    #    if os.path.exists(archivo):
    #        os.remove(archivo)

    def borrar_productos(self):
        confirmacion = QMessageBox.question(self, "Confirmar", "¿Seguro que querés borrar todos los productos del Excel?",
                                            QMessageBox.Yes | QMessageBox.No)
        archivo = "productos.xlsx"
        if confirmacion == QMessageBox.Yes:
            # Sobrescribe el archivo dejando solo los encabezados
            if os.path.exists(archivo):
                wb = openpyxl.Workbook()
                hoja = wb.active
                hoja.append(["Nombre", "Precio"])
                wb.save(archivo)
                self.lista_productos.clear()
                print("🧹 Productos borrados correctamente.")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ventana = ComanderaApp()
    ventana.show()
    sys.exit(app.exec_())